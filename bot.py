#!/usr/bin/env python3
"""
Semiconductor job scanner: pulls public listings from Greenhouse / Lever / Workday /
Ashby / Phenom-style / Mirafra WordPress career pages, scores titles for design / verification / performance
modeling, and tracks new vs already-seen and applied in SQLite.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
import re
import sqlite3
import warnings
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlparse, urljoin, urlunparse

# Fires on first urllib3 import (bundled with requests); filter before loading requests.
warnings.filterwarnings(
    "ignore",
    message=r"urllib3 v2 only supports OpenSSL.*",
    category=Warning,
)

import requests
import yaml
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "jobs.db"
COMPANIES_PATH = ROOT / "companies.yaml"
KEYWORDS_PATH = ROOT / "keywords.yaml"
LOGS_DIR = ROOT / "daily_logs"
EXCEL_TRACKER_PATH = ROOT / "jobs_tracker.xlsx"
USER_AGENT = "SemiconductorJobBot/1.0 (+local; job search)"

# JobSpy → Indeed GraphQL `dateOnIndeed` window (hours). Tighter than default 168h to reduce stale index rows.
JOBSPY_INDEED_DATE_ON_INDEED_HOURS = 72
# For Indeed viewjob links: HTTP-sniff page if posting date is missing or at least this many days old.
JOBSPY_INDEED_EXPIRY_PROBE_MIN_AGE_DAYS = 2

_EXCEL_JOB_ID_JR_RE = re.compile(r"JR[-_]?\d+", re.IGNORECASE)

_INDEED_EXPIRED_PAGE_MARKERS: tuple[str, ...] = (
    "this job has expired",
    "job has expired",
    "this job is no longer available",
    "no longer accepting applications",
    "no longer accepting new applications",
    "position has been filled",
    "this listing is no longer available",
    "job listing is no longer available",
)


@dataclass
class NormalizedJob:
    source_key: str
    external_id: str
    title: str
    company_name: str
    url: str
    location: str | None
    posted_at: str | None
    body: str


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS jobs (
            source_key TEXT NOT NULL,
            external_id TEXT NOT NULL,
            title TEXT NOT NULL,
            company_name TEXT NOT NULL,
            url TEXT NOT NULL UNIQUE,
            location TEXT,
            posted_at TEXT,
            first_seen_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            applied INTEGER NOT NULL DEFAULT 0,
            relevance_score REAL NOT NULL,
            tracks TEXT NOT NULL,
            PRIMARY KEY (source_key, external_id)
        );
        CREATE INDEX IF NOT EXISTS idx_jobs_applied ON jobs(applied);
        CREATE INDEX IF NOT EXISTS idx_jobs_first_seen ON jobs(first_seen_at);
        """
    )
    conn.commit()


def _session() -> requests.Session:
    s = requests.Session()
    if os.environ.get("SEMICONDUCTOR_JOB_BOT_NO_PROXY", "").strip() in ("1", "true", "yes"):
        s.trust_env = False
    return s


_SESSION = _session()


def http_get(url: str, timeout: float = 30.0) -> dict[str, Any] | list[Any]:
    r = _SESSION.get(
        url,
        headers={"User-Agent": USER_AGENT},
        timeout=timeout,
    )
    r.raise_for_status()
    return r.json()


ASHBY_JOB_BOARD_QUERY = (
    "query ApiJobBoardWithTeams($organizationHostedJobsPageName: String!) "
    "{ jobBoard: jobBoardWithTeams(organizationHostedJobsPageName: $organizationHostedJobsPageName) "
    "{ jobPostings { id title locationName isRemote externalLink updatedAt } } }"
)


def _workday_referer_from_tenant_url(tenant_url: str) -> str | None:
    """Build Referer like https://tenant/SiteName/ from /wday/cxs/.../SiteName/jobs."""
    parsed = urlparse(tenant_url)
    if not parsed.scheme or not parsed.netloc:
        return None
    segments = [s for s in parsed.path.split("/") if s]
    if len(segments) >= 2 and segments[-1] == "jobs":
        site = segments[-2]
        return urlunparse((parsed.scheme, parsed.netloc, f"/{site}/", "", "", ""))
    return urlunparse((parsed.scheme, parsed.netloc, "/", "", "", ""))


def _workday_site_from_tenant_url(tenant_url: str) -> str | None:
    parsed = urlparse(tenant_url)
    segments = [s for s in parsed.path.split("/") if s]
    if len(segments) >= 2 and segments[-1] == "jobs":
        return segments[-2]
    return None


def _workday_job_detail_api_url(tenant_url: str, external_path: str) -> str | None:
    """CXS single-posting JSON: same host as tenant_url, path .../SiteName + externalPath (e.g. /job/US-CA-...)."""
    ep = (external_path or "").strip()
    if not ep.startswith("/job/"):
        return None
    parsed = urlparse(tenant_url)
    path = parsed.path.rstrip("/")
    if path.endswith("/jobs"):
        base_path = path[: -len("/jobs")]
    else:
        base_path = path
    full_path = base_path.rstrip("/") + ep
    return urlunparse((parsed.scheme, parsed.netloc, full_path, "", "", ""))


_WORKDAY_ENRICH_PRIORITY_INTERN = re.compile(
    r"\b(intern|internship|co[-\s]?op)\b",
    re.IGNORECASE,
)
_WORKDAY_ENRICH_PRIORITY_GRAD = re.compile(
    r"\b(new\s+college\s+grad|new\s+grad|ncg|graduate)\b",
    re.IGNORECASE,
)
_WORKDAY_ENRICH_PRIORITY_HW = re.compile(
    r"\b(hardware|asic|rtl|verification|silicon|soc\b|gpu|cpu|design\s+verification|"
    r"dv\b|vlsi|fpga|emulation|modelling|modeling|architect|modem|chip)\b",
    re.IGNORECASE,
)


def _workday_enrich_priority_key(job: NormalizedJob) -> tuple[int, str]:
    t = job.title or ""
    if _WORKDAY_ENRICH_PRIORITY_INTERN.search(t):
        return (0, t.lower())
    if _WORKDAY_ENRICH_PRIORITY_GRAD.search(t):
        return (1, t.lower())
    if _WORKDAY_ENRICH_PRIORITY_HW.search(t):
        return (2, t.lower())
    return (3, t.lower())


def _workday_job_description_to_text(html: str) -> str:
    if not html or not str(html).strip():
        return ""
    soup = BeautifulSoup(str(html), "html.parser")
    return soup.get_text("\n", strip=True)


def _workday_enrich_job_bodies(
    tenant_url: str,
    portal_base: str,
    jobs: list[NormalizedJob],
    *,
    max_details: int,
) -> None:
    """
    Workday list responses usually omit jobDescription; fetch CXS detail JSON so scoring can use JD text.
    Stops after max_details successful fetches (plus one sleep per attempt).
    """
    if max_details <= 0 or not jobs:
        return
    parsed_tenant = urlparse(tenant_url)
    origin = urlunparse((parsed_tenant.scheme, parsed_tenant.netloc, "", "", "", ""))
    site = _workday_site_from_tenant_url(tenant_url)
    base = portal_base.rstrip("/")
    base_headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Accept-Language": "en-US",
        "Origin": origin,
    }
    hdr_prime = {**base_headers, "Content-Type": "application/json"}
    referer_candidates = [
        _workday_referer_from_tenant_url(tenant_url) or f"{origin}/",
        f"{origin}/{site}/jobs" if site else "",
        f"{base}/jobs",
        f"{base}/",
        f"{origin}/",
    ]
    referer_candidates = list(dict.fromkeys(r for r in referer_candidates if r))
    _workday_prime_session(origin, base, site, hdr_prime)

    prioritized = sorted(jobs, key=_workday_enrich_priority_key)
    got = 0
    for job in prioritized:
        if got >= max_details:
            break
        api_url = _workday_job_detail_api_url(tenant_url, job.external_id)
        if not api_url:
            continue
        desc_plain = ""
        for ref in referer_candidates:
            hdrs = {**base_headers, "Referer": ref}
            try:
                r = _SESSION.get(api_url, headers=hdrs, timeout=28.0)
                if r.status_code != 200:
                    continue
                payload = r.json()
            except (requests.RequestException, ValueError):
                continue
            if not isinstance(payload, dict):
                continue
            info = payload.get("jobPostingInfo")
            if not isinstance(info, dict):
                continue
            jd = info.get("jobDescription")
            if isinstance(jd, str) and jd.strip():
                desc_plain = _workday_job_description_to_text(jd)
                break
        time.sleep(0.18)
        if desc_plain:
            job.body = desc_plain[:65000]
            got += 1


def _workday_prime_session(origin: str, portal_base: str, site: str | None, base_headers: dict[str, str]) -> None:
    """Load HTML career pages so Workday/WAF may set cookies before CXS POST."""
    html_h = {
        **base_headers,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    del html_h["Content-Type"]
    pb = portal_base.rstrip("/")
    for url in (
        f"{origin}/",
        pb + "/",
        pb + "/jobs",
        f"{origin}/{site}/jobs" if site else "",
        f"{origin}/{site}/" if site else "",
    ):
        if not url or url.endswith("//"):
            continue
        try:
            _SESSION.get(url, headers=html_h, timeout=25.0)
        except requests.RequestException:
            pass


def _workday_post_bodies(page_limit: int, offset: int) -> list[dict[str, Any]]:
    """Tenants differ: some reject empty appliedFacets or require minimal keys only."""
    return [
        {
            "appliedFacets": {},
            "limit": page_limit,
            "offset": offset,
            "searchText": "",
            "sort": [{"field": "postedOn", "descending": True}],
        },
        {"appliedFacets": {}, "limit": page_limit, "offset": offset, "searchText": ""},
        {"limit": page_limit, "offset": offset, "searchText": ""},
    ]


def _workday_pull_all_pages(
    tenant_url: str,
    portal_base: str,
    company_name: str,
    *,
    silent_fail: bool = False,
) -> list[NormalizedJob] | None:
    """
    Pull all pages from one CXS endpoint. Returns None if the first page cannot be loaded (422/4xx).
    Returns [] if the endpoint works but has zero postings.
    """
    parsed_tenant = urlparse(tenant_url)
    host = parsed_tenant.netloc or "unknown"
    source_key = f"workday:{host}"
    origin = urlunparse((parsed_tenant.scheme, parsed_tenant.netloc, "", "", "", ""))
    referer = _workday_referer_from_tenant_url(tenant_url) or f"{origin}/"
    site = _workday_site_from_tenant_url(tenant_url)
    base = portal_base.rstrip("/")

    base_headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Origin": origin,
        "Accept-Language": "en-US",
    }

    referer_candidates = [
        referer,
        f"{origin}/{site}/jobs" if site else "",
        f"{base}/jobs",
        f"{base}/",
        f"{origin}/",
    ]
    referer_candidates = list(dict.fromkeys(r for r in referer_candidates if r))

    _workday_prime_session(origin, base, site, base_headers)

    page_limit = 20
    out: list[NormalizedJob] = []
    seen_paths: set[str] = set()
    offset = 0

    while True:
        time.sleep(1.5 if offset == 0 else 0.35)
        bodies = _workday_post_bodies(page_limit, offset)
        data: dict[str, Any] | None = None
        last_err: str = ""
        for ref in referer_candidates:
            hdrs = {**base_headers, "Referer": ref}
            for body in bodies:
                try:
                    r = _SESSION.post(tenant_url, json=body, headers=hdrs, timeout=30.0)
                except requests.RequestException as e:
                    last_err = str(e)
                    continue
                if r.status_code in (400, 422):
                    last_err = f"HTTP {r.status_code}"
                    time.sleep(0.2)
                    continue
                if r.status_code >= 400:
                    last_err = f"HTTP {r.status_code}"
                    break
                try:
                    parsed = r.json()
                except ValueError as e:
                    last_err = str(e)
                    continue
                if isinstance(parsed, dict):
                    data = parsed
                    break
            if data is not None:
                break

        if data is None:
            if offset == 0:
                if not silent_fail:
                    print(
                        f"[warn] Workday {company_name} ({urlparse(tenant_url).path}): {last_err}",
                        file=sys.stderr,
                    )
                return None
            break

        postings = data.get("jobPostings") or []
        if not isinstance(postings, list):
            break
        if not postings:
            break

        for j in postings:
            if not isinstance(j, dict):
                continue
            path = (j.get("externalPath") or "").strip()
            if path and path in seen_paths:
                continue
            if path:
                seen_paths.add(path)

            title = (j.get("title") or "").strip()
            if path.startswith("http://") or path.startswith("https://"):
                url = path
            else:
                if not path:
                    continue
                path_part = path if path.startswith("/") else f"/{path}"
                url = f"{base}{path_part}"
            ext_id = path if path else url
            loc_raw = j.get("locationsText")
            location = (str(loc_raw).strip() if loc_raw is not None else "") or None
            # Workday list API returns human strings, e.g. "Posted 20 Days Ago", not timestamps.
            posted = j.get("postedOn")
            posted_at = str(posted).strip() if posted is not None and str(posted).strip() else None
            out.append(
                NormalizedJob(
                    source_key=source_key,
                    external_id=str(ext_id),
                    title=title,
                    company_name=company_name,
                    url=url,
                    location=location,
                    posted_at=posted_at,
                    body="",
                )
            )

        total = data.get("total")
        if isinstance(total, int) and offset + len(postings) >= total:
            break
        if len(postings) < page_limit:
            break
        offset += page_limit

    return out


def _workday_alternate_pairs(raw: Any) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    if not isinstance(raw, list):
        return pairs
    for item in raw:
        if not isinstance(item, dict):
            continue
        tu = item.get("tenant_url")
        pb = item.get("portal_base")
        if isinstance(tu, str) and isinstance(pb, str) and tu.strip() and pb.strip():
            pairs.append((tu.strip(), pb.strip()))
    return pairs


def fetch_workday(
    tenant_url: str,
    portal_base: str,
    company_name: str,
    workday_alternate: Any = None,
    *,
    enrich_job_details: bool = False,
    enrich_max: int = 0,
) -> list[NormalizedJob]:
    candidates: list[tuple[str, str]] = [(tenant_url, portal_base)]
    candidates.extend(_workday_alternate_pairs(workday_alternate))
    seen_u: set[tuple[str, str]] = set()
    uniq: list[tuple[str, str]] = []
    for tu, pb in candidates:
        key = (tu, pb)
        if key not in seen_u:
            seen_u.add(key)
            uniq.append((tu, pb))
    n = len(uniq)
    for i, (tu, pb) in enumerate(uniq):
        got = _workday_pull_all_pages(
            tu, pb, company_name, silent_fail=n > 1 and i < n - 1
        )
        if got is not None:
            if enrich_job_details and enrich_max > 0:
                try:
                    em = int(enrich_max)
                except (TypeError, ValueError):
                    em = 0
                if em > 0:
                    _workday_enrich_job_bodies(tu, pb, got, max_details=em)
            return got
    return []


def _eightfold_response_looks_like_search(d: dict[str, Any]) -> bool:
    if _eightfold_flat_records(d):
        return True
    for k in ("total", "totalSize", "count", "positions", "records", "jobs", "results"):
        if k in d:
            return True
    inner = d.get("data")
    if isinstance(inner, dict):
        for k in ("positions", "records", "jobs", "results", "total", "totalSize"):
            if k in inner:
                return True
    return False


def _eightfold_flat_records(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    for key in ("positions", "records", "jobs", "results"):
        v = payload.get(key)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]
    inner = payload.get("data")
    if isinstance(inner, dict):
        for key in ("positions", "records", "jobs", "results"):
            v = inner.get(key)
            if isinstance(v, list):
                return [x for x in v if isinstance(x, dict)]
    return []


def _eightfold_str(d: dict[str, Any], *keys: str) -> str | None:
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def _eightfold_prime_origin(origin: str, domain: str, user_agent: str) -> None:
    """PCS often requires cookies from the careers shell before /api/careers/* accepts requests."""
    html_headers = {
        "User-Agent": user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US",
    }
    q = urlencode({"domain": domain})
    for path in (f"/careers?{q}", f"/careers/search?{q}", "/careers"):
        try:
            _SESSION.get(f"{origin.rstrip('/')}{path}", headers=html_headers, timeout=28.0)
        except requests.RequestException:
            pass


def fetch_eightfold(careers_host: str, domain: str, company_name: str) -> list[NormalizedJob]:
    """
    Eightfold PCS job search (Qualcomm and others). Public careers sites POST to
    /api/careers/v2/search/jobs on the same host as the careers UI.
    """
    raw = careers_host.strip()
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "https://" + raw
    origins: list[str] = []
    if "://" in raw:
        parsed = urlparse(raw)
        h = parsed.netloc or ""
        if h:
            origins.append(urlunparse((parsed.scheme or "https", h, "", "", "", "")))
    else:
        h0 = raw.split("/")[0]
        if h0:
            origins.append(f"https://{h0}")
    dom_l = domain.lower().strip()
    if dom_l == "qualcomm.com":
        # Prefer the same host as the careers UI so cookies + CORS align; fall back to eightfold.ai.
        for extra in (
            "https://careers.qualcomm.com",
            "https://qualcomm.eightfold.ai",
            "https://app.eightfold.ai",
        ):
            if extra not in origins:
                origins.append(extra)
        preferred = (
            "https://careers.qualcomm.com",
            "https://qualcomm.eightfold.ai",
            "https://app.eightfold.ai",
        )
        seen_p: set[str] = set()
        reordered: list[str] = []
        for o in preferred:
            if o in origins and o not in seen_p:
                reordered.append(o)
                seen_p.add(o)
        for o in origins:
            if o not in seen_p:
                reordered.append(o)
                seen_p.add(o)
        origins = reordered
    else:
        origins = list(dict.fromkeys(origins))
    if not origins:
        print(f"[warn] Eightfold {company_name}: invalid eightfold_host", file=sys.stderr)
        return []

    locked_origin: str | None = None
    host = urlparse(origins[0]).netloc or "unknown"
    source_key = f"eightfold:{host}"
    # Do not use /careers/v2/search/jobs — on careers.qualcomm.com it is not a real route (GET 404).
    search_paths = (
        "/api/careers/v2/search/jobs",
        "/api/careers/v3/talent/job/search",
    )
    base_headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Accept-Language": "en-US",
    }

    out: list[NormalizedJob] = []
    seen: set[str] = set()
    start = 0
    page_size = 30
    max_pages = 80
    primed_origins: set[str] = set()

    for _ in range(max_pages):
        time.sleep(1.2 if start == 0 else 0.4)
        body_variants: list[dict[str, Any]] = [
            {"domain": domain, "start": start, "limit": page_size},
            {
                "domain": domain,
                "start": start,
                "limit": page_size,
                "query": "",
                "location": [],
            },
        ]
        data: dict[str, Any] | None = None
        last_err: str | None = None
        try_origins = [locked_origin] if locked_origin else origins
        for origin in try_origins:
            if not origin:
                continue
            if origin not in primed_origins:
                _eightfold_prime_origin(origin, domain, base_headers["User-Agent"])
                primed_origins.add(origin)
            hdrs = {
                **base_headers,
                "Origin": origin,
                "Referer": f"{origin}/careers?domain={domain}",
            }
            hdrs_get = {
                "User-Agent": base_headers["User-Agent"],
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-US",
                "Origin": origin,
                "Referer": f"{origin}/careers?domain={domain}",
            }
            page_ok = False
            for path in search_paths:
                url = f"{origin}{path}"
                for body in body_variants:
                    try:
                        r = _SESSION.post(url, json=body, headers=hdrs, timeout=35.0)
                    except requests.RequestException as e:
                        last_err = str(e)
                        continue
                    if r.status_code == 404:
                        last_err = f"POST 404 {path}"
                        continue
                    if r.status_code in (401, 403):
                        last_err = f"POST HTTP {r.status_code} {origin}{path}"
                        continue
                    if r.status_code >= 400:
                        last_err = f"POST HTTP {r.status_code}"
                        continue
                    try:
                        parsed_json = r.json()
                    except ValueError:
                        last_err = "invalid JSON"
                        continue
                    if isinstance(parsed_json, dict) and _eightfold_response_looks_like_search(
                        parsed_json
                    ):
                        data = parsed_json
                        locked_origin = origin
                        page_ok = True
                        break
                if page_ok:
                    break
                if path.startswith("/api/"):
                    try:
                        r = _SESSION.get(
                            url,
                            params={
                                "domain": domain,
                                "start": str(start),
                                "limit": str(page_size),
                            },
                            headers=hdrs_get,
                            timeout=35.0,
                        )
                    except requests.RequestException as e:
                        last_err = str(e)
                        continue
                    if r.status_code in (401, 403, 404):
                        last_err = f"GET HTTP {r.status_code} {origin}{path}"
                        continue
                    if r.status_code >= 400:
                        continue
                    try:
                        parsed_json = r.json()
                    except ValueError:
                        continue
                    if isinstance(parsed_json, dict) and _eightfold_response_looks_like_search(
                        parsed_json
                    ):
                        data = parsed_json
                        locked_origin = origin
                        page_ok = True
                        break
            if page_ok:
                break

        if data is None:
            if start == 0:
                print(f"[warn] Eightfold {company_name}: {last_err or 'no working search path'}", file=sys.stderr)
            break

        records = _eightfold_flat_records(data)
        if not records:
            break

        for rec in records:
            title = _eightfold_str(rec, "name", "title", "display_name", "job_title") or ""
            jid = _eightfold_str(rec, "id", "job_id", "req_id", "number") or title or ""
            url_j = _eightfold_str(
                rec,
                "t_share_url",
                "canonical_position_url",
                "url",
                "apply_url",
                "job_url",
            )
            if not url_j or not title:
                continue
            loc: str | None = None
            jl = rec.get("job_location")
            if isinstance(jl, str) and jl.strip():
                loc = jl.strip()
            elif isinstance(jl, dict):
                loc = _eightfold_str(jl, "name", "city", "display_name", "location_name")
            locs = rec.get("locations")
            if not loc and isinstance(locs, list) and locs:
                first = locs[0]
                if isinstance(first, dict):
                    loc = _eightfold_str(first, "name", "city", "location_name", "display_name")
                elif isinstance(first, str) and first.strip():
                    loc = first.strip()

            posted_raw = rec.get("t_create_time") or rec.get("create_time") or rec.get("posted_date")
            posted_at = str(posted_raw).strip() if posted_raw is not None else None

            key = f"{jid}|{url_j}"
            if key in seen:
                continue
            seen.add(key)
            out.append(
                NormalizedJob(
                    source_key=source_key,
                    external_id=str(jid),
                    title=title,
                    company_name=company_name,
                    url=url_j,
                    location=loc,
                    posted_at=posted_at,
                    body="",
                )
            )

        total = data.get("total") or data.get("totalSize") or data.get("count")
        n = len(records)
        if isinstance(total, int) and start + n >= total:
            break
        if n < page_size:
            break
        start += page_size

    return out

def fetch_ashby(org: str, company_name: str) -> list[NormalizedJob]:
    source_key = f"ashby:{org}"
    url = "https://jobs.ashbyhq.com/api/non-user-graphql?op=ApiJobBoardWithTeams"
    payload = {
        "operationName": "ApiJobBoardWithTeams",
        "variables": {"organizationHostedJobsPageName": org},
        "query": ASHBY_JOB_BOARD_QUERY,
    }
    headers = {
        "User-Agent": USER_AGENT,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    r = _SESSION.post(url, json=payload, headers=headers, timeout=30.0)
    r.raise_for_status()
    data = r.json()
    board = (data.get("data") or {}).get("jobBoard") or {}
    jobs = board.get("jobPostings") or []
    if not isinstance(jobs, list):
        return []

    out: list[NormalizedJob] = []
    for j in jobs:
        if not isinstance(j, dict):
            continue
        jid = j.get("id")
        if jid is None:
            continue
        jid_str = str(jid)
        title = (j.get("title") or "").strip()
        ext_link = j.get("externalLink")
        if ext_link:
            job_url = str(ext_link).strip()
        else:
            job_url = f"https://jobs.ashbyhq.com/{org}/{jid_str}"
        loc_name = j.get("locationName")
        location = (str(loc_name).strip() if loc_name is not None else "") or None
        upd = j.get("updatedAt")
        posted_at = str(upd) if upd is not None else None
        out.append(
            NormalizedJob(
                source_key=source_key,
                external_id=jid_str,
                title=title,
                company_name=company_name,
                url=job_url,
                location=location,
                posted_at=posted_at,
                body="",
            )
        )
    return out


def _phenom_listing_url_page(base_url: str, page: int, page_query: str = "pg") -> str:
    pq = page_query.strip() or "pg"
    if page <= 1:
        return base_url
    p = urlparse(base_url)
    items = [(k, v) for k, v in parse_qsl(p.query, keep_blank_values=True) if k != pq]
    items.append((pq, str(page)))
    new_q = urlencode(items)
    return urlunparse((p.scheme, p.netloc, p.path, p.params, new_q, p.fragment))


PHENOM_HTML_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US",
}


def fetch_phenom(
    listing_url: str,
    company_name: str,
    max_pages: int = 15,
    page_query: str = "pg",
) -> list[NormalizedJob]:
    """
    Parse job rows from a Phenom-style search HTML page (e.g. careers.arm.com/search-jobs).
    Pagination uses ?pg= (default) or another query name (e.g. Renesas uses ?page=).
    """
    host = urlparse(listing_url).netloc or "unknown"
    source_key = f"phenom:{host}"
    seen_urls: set[str] = set()
    out: list[NormalizedJob] = []
    pq = (page_query or "pg").strip() or "pg"

    for page in range(1, max(1, max_pages) + 1):
        url = _phenom_listing_url_page(listing_url, page, page_query=pq)
        time.sleep(0.35)
        try:
            r = _SESSION.get(url, headers=PHENOM_HTML_HEADERS, timeout=30.0)
            r.raise_for_status()
        except requests.RequestException as e:
            if page == 1:
                raise
            print(f"[warn] Phenom {company_name}: stopped at page {page} ({e})", file=sys.stderr)
            break

        soup = BeautifulSoup(r.text, "html.parser")
        page_new = 0
        for a in soup.select('a[href*="/job/"]'):
            href = (a.get("href") or "").strip()
            if not href:
                continue
            full = urljoin(listing_url, href)
            if full in seen_urls or "/job/" not in full:
                continue
            title = a.get_text(strip=True)
            if not title:
                continue
            seen_urls.add(full)
            page_new += 1
            path_parts = [x for x in urlparse(full).path.split("/") if x]
            ext_id = path_parts[-1] if path_parts else full
            location: str | None = None
            li = a.find_parent("li")
            if li:
                raw = li.get_text(" ", strip=True)
                if raw.startswith(title):
                    rest = raw[len(title) :].strip()
                    if rest:
                        location = rest[:280]
            out.append(
                NormalizedJob(
                    source_key=source_key,
                    external_id=str(ext_id),
                    title=title,
                    company_name=company_name,
                    url=full,
                    location=location,
                    posted_at=None,
                    body="",
                )
            )

        if page_new == 0:
            break

    return out


def fetch_jobvite(site_slug: str, company_name: str) -> list[NormalizedJob]:
    """Scrape Jobvite-hosted list pages (e.g. jobs.jobvite.com/cirruslogic/jobs)."""
    slug = site_slug.strip().strip("/")
    source_key = f"jobvite:{slug}"
    root = f"https://jobs.jobvite.com/{slug}"
    list_urls = [f"{root}/jobs", f"{root}/join", root]
    seen_urls: set[str] = set()
    out: list[NormalizedJob] = []

    for list_url in list_urls:
        time.sleep(0.4)
        try:
            r = _SESSION.get(list_url, headers=PHENOM_HTML_HEADERS, timeout=30.0)
            r.raise_for_status()
        except requests.RequestException:
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        page_new = 0
        for a in soup.find_all("a", href=True):
            href = (a.get("href") or "").strip()
            if "/job/" not in href.lower():
                continue
            full = urljoin(list_url, href)
            if full in seen_urls or "jobvite" not in full.lower():
                continue
            title = a.get_text(strip=True)
            if not title or len(title) < 3:
                continue
            seen_urls.add(full)
            page_new += 1
            path_parts = [x for x in urlparse(full).path.split("/") if x]
            ext_id = path_parts[-1] if path_parts else full
            out.append(
                NormalizedJob(
                    source_key=source_key,
                    external_id=str(ext_id),
                    title=title,
                    company_name=company_name,
                    url=full,
                    location=None,
                    posted_at=None,
                    body="",
                )
            )
        if page_new > 0:
            break

    if not out:
        print(f"[warn] Jobvite {company_name}: no job links at {root}", file=sys.stderr)
    return out


def fetch_mirafra(careers_url: str, company_name: str) -> list[NormalizedJob]:
    """
    Mirafra lists roles as static cards on mirafra.com/career/ (no Greenhouse/Lever URL per req).
    external_id is the modal anchor id when present (unique per row).
    """
    base = (careers_url or "").strip() or "https://mirafra.com/career/"
    if not base.startswith("http://") and not base.startswith("https://"):
        base = "https://" + base
    source_key = "mirafra:career"
    try:
        r = _SESSION.get(base, headers=PHENOM_HTML_HEADERS, timeout=35.0)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"[warn] Mirafra {company_name}: {e}", file=sys.stderr)
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    out: list[NormalizedJob] = []
    seen: set[str] = set()
    for div in soup.select("div.row.res_content"):
        h3 = div.select_one(".postion-info h3")
        if not h3:
            continue
        title = h3.get_text(strip=True)
        if not title or len(title) < 2:
            continue
        exp_el = div.select_one(".postion-info p")
        exp_line = exp_el.get_text(strip=True) if exp_el else ""
        btn = div.select_one("button.btn-career[data-target]")
        tid = (btn.get("data-target") or "").lstrip("#").strip() if btn else ""
        if not tid:
            tid = hashlib.sha256(f"{title}\0{exp_line}".encode()).hexdigest()[:24]
        if tid in seen:
            continue
        seen.add(tid)
        body = exp_line
        job_url = f"{base.rstrip('/')}/#{tid}"
        out.append(
            NormalizedJob(
                source_key=source_key,
                external_id=tid,
                title=title,
                company_name=company_name,
                url=job_url,
                location=None,
                posted_at=None,
                body=body,
            )
        )

    if not out:
        print(f"[warn] Mirafra {company_name}: no job rows at {base}", file=sys.stderr)
    return out


def fetch_smartrecruiters(company_identifier: str, company_name: str) -> list[NormalizedJob]:
    """
    Public Posting API v1 (no key) when the tenant allows it.
    https://developers.smartrecruiters.com/reference/listpostings-1
    """
    cid = company_identifier.strip()
    source_key = f"smartrecruiters:{cid}"
    base_api = f"https://api.smartrecruiters.com/v1/companies/{cid}/postings"
    out: list[NormalizedJob] = []
    offset = 0
    limit = 100
    max_pages = 50

    for _ in range(max_pages):
        time.sleep(0.35)
        try:
            r = _SESSION.get(
                base_api,
                params={"offset": offset, "limit": limit},
                headers={
                    "User-Agent": PHENOM_HTML_HEADERS["User-Agent"],
                    "Accept": "application/json",
                },
                timeout=30.0,
            )
            if r.status_code in (401, 403, 404):
                if offset == 0:
                    print(
                        f"[warn] SmartRecruiters {company_name}: HTTP {r.status_code} "
                        f"(try ats: phenom + listing_url if listings use /job/ links)",
                        file=sys.stderr,
                    )
                break
            r.raise_for_status()
            data = r.json()
        except requests.RequestException as e:
            if offset == 0:
                print(f"[warn] SmartRecruiters {company_name}: {e}", file=sys.stderr)
            break
        except ValueError:
            break

        content: Any = None
        if isinstance(data, dict):
            content = data.get("content")
            if not isinstance(content, list):
                content = data.get("elements")
        if not isinstance(content, list):
            break
        if not content:
            break

        for posting in content:
            if not isinstance(posting, dict):
                continue
            pid = posting.get("id") or posting.get("uuid")
            title = (posting.get("name") or posting.get("title") or "").strip()
            if pid is None or not title:
                continue
            loc: str | None = None
            loc_obj = posting.get("location")
            if isinstance(loc_obj, dict):
                parts = [
                    loc_obj.get("city"),
                    loc_obj.get("region"),
                    loc_obj.get("country"),
                ]
                loc = ", ".join(str(p) for p in parts if p) or None
            job_url = (
                posting.get("applyUrl")
                or posting.get("leadsDirectApplyUrl")
                or posting.get("referralUrl")
                or posting.get("jobAdUrl")
            )
            if not job_url or not isinstance(job_url, str):
                job_url = f"https://jobs.smartrecruiters.com/{cid}/{pid}"
            released = posting.get("releasedDate") or posting.get("updatedOn")
            posted_at = str(released) if released is not None else None
            out.append(
                NormalizedJob(
                    source_key=source_key,
                    external_id=str(pid),
                    title=title,
                    company_name=company_name,
                    url=job_url.strip(),
                    location=loc,
                    posted_at=posted_at,
                    body="",
                )
            )

        total = data.get("totalFound") if isinstance(data, dict) else None
        if isinstance(total, int) and offset + len(content) >= total:
            break
        if len(content) < limit:
            break
        offset += limit

    return out


def _indeed_title_merge_key(title: str) -> str:
    """
    Normalize title so Indeed rows cluster when they are the same role.
    Uses the full normalized title when there is no strong suffix; otherwise uses the
    substring after the last ' - ' if it is long enough. That folds variants like
    '…Internships (US) - Machine Learning…' and '…Early Career (US) - Machine Learning…'
    into one key without merging unrelated short-titled posts.
    """
    raw = " ".join((title or "").strip().split()).casefold()
    if " - " in raw:
        tail = raw.rsplit(" - ", 1)[-1].strip()
        if len(tail) >= 20:
            return tail
    return raw


def _merge_jobspy_indeed_by_title(jobs: list[NormalizedJob]) -> list[NormalizedJob]:
    """
    Indeed / JobSpy often returns the same role as multiple rows (different metros)
    or as separate listings that differ only in wording before a shared suffix
    (e.g. Internships vs Early Career). Same merge key (per source_key): combine
    locations, keep one primary URL, list alternates in body (one URL per line).
    """
    groups: dict[tuple[str, str], list[NormalizedJob]] = {}
    for j in jobs:
        key = (j.source_key, _indeed_title_merge_key(j.title))
        groups.setdefault(key, []).append(j)
    out: list[NormalizedJob] = []
    for group in groups.values():
        if len(group) == 1:
            out.append(group[0])
            continue
        group.sort(key=lambda x: ((x.url or ""), (x.external_id or "")))
        locs_ordered: list[str] = []
        seen_loc: set[str] = set()
        for j in group:
            if not j.location or not str(j.location).strip():
                continue
            loc = str(j.location).strip()
            lk = loc.casefold()
            if lk not in seen_loc:
                seen_loc.add(lk)
                locs_ordered.append(loc)
        merged_location = " | ".join(sorted(locs_ordered, key=str.casefold)) if locs_ordered else None
        urls: list[str] = []
        seen_u: set[str] = set()
        for j in group:
            u = (j.url or "").strip()
            if u and u not in seen_u:
                seen_u.add(u)
                urls.append(u)
        primary_url = urls[0]
        alt_urls = urls[1:]
        posted_candidates = [j.posted_at for j in group if j.posted_at]
        merged_posted = max(posted_candidates) if posted_candidates else None
        ref = group[0]
        title_ref = max(group, key=lambda j: (len((j.title or "").strip()), j.title or ""))
        mk = _indeed_title_merge_key(title_ref.title)
        h = hashlib.sha256(
            f"{mk}\0{title_ref.company_name.strip().casefold()}".encode()
        ).hexdigest()[:20]
        ext_id = f"indeed-merge:{h}"
        body = "\n".join(alt_urls) if alt_urls else ""
        out.append(
            NormalizedJob(
                source_key=ref.source_key,
                external_id=ext_id,
                title=title_ref.title.strip(),
                company_name=ref.company_name,
                url=primary_url,
                location=merged_location,
                posted_at=merged_posted,
                body=body,
            )
        )
    out.sort(key=lambda j: (j.company_name.casefold(), j.title.casefold()))
    return out


def _indeed_jobspy_needs_expiry_probe(posted_at: str | None) -> bool:
    if not posted_at or not str(posted_at).strip():
        return True
    try:
        d = datetime.strptime(str(posted_at).strip()[:10], "%Y-%m-%d").date()
        age = (datetime.now(timezone.utc).date() - d).days
        return age >= JOBSPY_INDEED_EXPIRY_PROBE_MIN_AGE_DAYS
    except ValueError:
        return True


def _indeed_viewjob_page_looks_active(url: str, timeout: float = 14.0) -> bool:
    """Best-effort: Indeed still serves some filled/expired reqs in search; drop obvious expiry HTML."""
    if os.environ.get("SEMICONDUCTOR_JOB_BOT_SKIP_INDEED_EXPIRY_CHECK", "").strip().lower() in (
        "1",
        "true",
        "yes",
    ):
        return True
    try:
        with _SESSION.get(
            url,
            headers={"User-Agent": USER_AGENT},
            timeout=timeout,
            stream=True,
        ) as r:
            if r.status_code >= 500:
                return True
            buf = bytearray()
            for chunk in r.iter_content(chunk_size=32768):
                if not chunk:
                    break
                buf.extend(chunk)
                if len(buf) >= 131072:
                    break
        low = bytes(buf).decode("utf-8", errors="ignore").lower()
        return not any(m in low for m in _INDEED_EXPIRED_PAGE_MARKERS)
    except (OSError, requests.RequestException):
        return True


def _drop_expired_indeed_jobspy_rows(jobs: list[NormalizedJob]) -> list[NormalizedJob]:
    kept: list[NormalizedJob] = []
    for j in jobs:
        u = (j.url or "").strip()
        if "indeed.com" not in u or "viewjob" not in u:
            kept.append(j)
            continue
        if not _indeed_jobspy_needs_expiry_probe(j.posted_at):
            kept.append(j)
            continue
        ok = _indeed_viewjob_page_looks_active(u)
        time.sleep(0.06)
        if ok:
            kept.append(j)
    return kept


def fetch_jobspy(search_queries: list[str], company_name: str) -> list[NormalizedJob]:
    try:
        from jobspy import scrape_jobs
    except ImportError:
        print(
            f"[warn] JobSpy not installed (Python 3.10+). "
            f"Run ./install.sh or: pip install -r requirements.txt. Skipping {company_name}.",
            file=sys.stderr,
        )
        return []

    source_key = f"jobspy:{company_name}"
    seen_ids: set[str] = set()
    out: list[NormalizedJob] = []

    for query in search_queries:
        q = str(query).strip()
        if not q:
            continue
        for job_type_filter in ["internship", "fulltime", None]:
            try:
                time.sleep(2.5)
                kwargs = dict(
                    site_name=["indeed"],
                    search_term=q,
                    location="United States",
                    results_wanted=15,
                    hours_old=JOBSPY_INDEED_DATE_ON_INDEED_HOURS,
                    country_indeed="USA",
                )
                if job_type_filter is not None:
                    kwargs["job_type"] = job_type_filter
                jobs_df = scrape_jobs(**kwargs)
                if jobs_df is None or jobs_df.empty:
                    continue
                for _, row in jobs_df.iterrows():
                    jid = str(row.get("id") or row.get("job_url") or "")
                    if not jid or jid in seen_ids:
                        continue
                    seen_ids.add(jid)
                    title = str(row.get("title") or "").strip()
                    location = str(row.get("location") or "").strip()
                    url = str(row.get("job_url") or "").strip()
                    posted = str(row.get("date_posted") or "").strip()
                    if not title or not url:
                        continue
                    out.append(
                        NormalizedJob(
                            source_key=source_key,
                            external_id=jid,
                            title=title,
                            company_name=company_name,
                            url=url,
                            location=location if location else None,
                            posted_at=posted if posted else None,
                            body="",
                        )
                    )
            except Exception as e:
                print(
                    f"[warn] JobSpy {company_name} query='{q}' "
                    f"job_type={job_type_filter!r}: {e}",
                    file=sys.stderr,
                )

    if not out:
        print(f"[warn] JobSpy {company_name}: no jobs matched queries", file=sys.stderr)
        return []
    out = _drop_expired_indeed_jobspy_rows(out)
    return _merge_jobspy_indeed_by_title(out)


def fetch_greenhouse(board: str, company_name: str) -> list[NormalizedJob]:
    source_key = f"greenhouse:{board}"
    url = f"https://boards-api.greenhouse.io/v1/boards/{board}/jobs?content=true"
    data = http_get(url)
    jobs = data.get("jobs") or []
    out: list[NormalizedJob] = []
    for j in jobs:
        jid = str(j.get("id", ""))
        title = (j.get("title") or "").strip()
        loc = j.get("location") or {}
        loc_name = loc.get("name") if isinstance(loc, dict) else None
        content = j.get("content") or ""
        if isinstance(content, str):
            body = content
        else:
            body = json.dumps(content) if content else ""
        out.append(
            NormalizedJob(
                source_key=source_key,
                external_id=jid,
                title=title,
                company_name=company_name,
                url=(j.get("absolute_url") or "").strip(),
                location=loc_name,
                posted_at=j.get("updated_at") or j.get("first_published"),
                body=body,
            )
        )
    return out


def fetch_lever(site: str, company_name: str) -> list[NormalizedJob]:
    source_key = f"lever:{site}"
    url = f"https://api.lever.co/v0/postings/{site}?mode=json"
    data = http_get(url)
    if not isinstance(data, list):
        return []
    out: list[NormalizedJob] = []
    for j in data:
        jid = str(j.get("id", ""))
        title = (j.get("text") or "").strip()
        locs = j.get("categories", {}).get("location", "")
        if isinstance(locs, str):
            loc_name = locs or None
        else:
            loc_name = None
        desc = (j.get("description") or "") or (j.get("descriptionPlain") or "")
        hosted = (j.get("hostedUrl") or j.get("applyUrl") or "").strip()
        out.append(
            NormalizedJob(
                source_key=source_key,
                external_id=jid,
                title=title,
                company_name=company_name,
                url=hosted,
                location=loc_name,
                posted_at=j.get("createdAt") or j.get("updatedAt"),
                body=desc if isinstance(desc, str) else "",
            )
        )
    return out


def skip_performance_track_company_set(raw: dict[str, Any]) -> frozenset[str]:
    xs = raw.get("skip_performance_track_for_companies")
    if not isinstance(xs, list):
        return frozenset()
    return frozenset(x.strip().casefold() for x in xs if isinstance(x, str) and x.strip())


def compile_keyword_config(
    raw: dict[str, Any],
) -> tuple[dict[str, list[re.Pattern[str]]], dict[str, list[tuple[re.Pattern[str], ...]]], float]:
    """
    Returns (exact_patterns, token_group_patterns, min_score).
    exact_patterns — single compiled regex per keyword phrase (original behavior).
    token_group_patterns — for multi-word phrases, a tuple of per-token patterns;
                           all tokens must match for the group to count as one hit.
    """
    exact: dict[str, list[re.Pattern[str]]] = {}
    token_groups: dict[str, list[tuple[re.Pattern[str], ...]]] = {}

    for key in ("design", "verification", "performance_modeling"):
        phrases = raw.get(key) or []
        ep: list[re.Pattern[str]] = []
        tg: list[tuple[re.Pattern[str], ...]] = []
        for p in phrases:
            if not isinstance(p, str) or not p.strip():
                continue
            phrase = p.strip().lower()
            esc = re.escape(phrase)
            ep.append(re.compile(rf"\b{esc}\b", re.IGNORECASE))
            tokens = phrase.split()
            if len(tokens) > 1:
                tg.append(
                    tuple(
                        re.compile(rf"\b{re.escape(t)}\b", re.IGNORECASE)
                        for t in tokens
                    )
                )
        exact[key] = ep
        token_groups[key] = tg

    min_score = float(raw.get("min_score", 1))
    return exact, token_groups, min_score


def compile_exclude_title_patterns(raw: dict[str, Any]) -> list[re.Pattern[str]]:
    """Regexes from keywords.yaml `exclude_title_regex`; any match drops the job."""
    patterns: list[re.Pattern[str]] = []
    xs: Any = raw.get("exclude_title_regex")
    if isinstance(xs, str) and xs.strip():
        xs = [xs]
    if not isinstance(xs, list):
        return patterns
    for item in xs:
        if not isinstance(item, str) or not item.strip():
            continue
        try:
            patterns.append(re.compile(item.strip(), re.IGNORECASE))
        except re.error:
            print(f"[warn] invalid exclude_title_regex: {item!r}", file=sys.stderr)
    return patterns


def job_title_excluded(title: str, patterns: list[re.Pattern[str]]) -> bool:
    if not patterns:
        return False
    t = title or ""
    return any(p.search(t) for p in patterns)


def score_job(
    text: str,
    tracks: dict[str, list[re.Pattern[str]]],
    *,
    token_groups: dict[str, list[tuple[re.Pattern[str], ...]]] | None = None,
    company_name: str = "",
    skip_performance_companies: frozenset[str] | None = None,
) -> tuple[float, list[str]]:
    hay = text or ""
    matched_tracks: list[str] = []
    total = 0.0
    cn = (company_name or "").strip().casefold()
    skip_perf = (
        bool(skip_performance_companies)
        and cn in (skip_performance_companies or frozenset())
    )
    tg = token_groups or {}
    for name, pats in tracks.items():
        if skip_perf and name == "performance_modeling":
            continue
        track_hits = 0
        matched_phrases: set[str] = set()
        # Exact phrase matches (original behavior)
        for pat in pats:
            m = pat.search(hay)
            if m:
                phrase = m.group(0).lower()
                if phrase not in matched_phrases:
                    matched_phrases.add(phrase)
                    track_hits += 1
        # Token-group matches: all tokens of a multi-word phrase present anywhere
        for token_tuple in tg.get(name, []):
            if all(t.search(hay) for t in token_tuple):
                # Use pattern source as dedup key
                key = "|".join(t.pattern for t in token_tuple)
                if key not in matched_phrases:
                    matched_phrases.add(key)
                    track_hits += 1
        if track_hits:
            matched_tracks.append(name)
            total += min(3.0, 1.0 + 0.25 * (track_hits - 1))
    return total, matched_tracks


DEFAULT_US_LOCATION_INDICATORS: tuple[str, ...] = (
    "United States",
    "United States of America",
    "USA",
    "U.S.A.",
    "U.S.",
    "Remote - United States",
    "Remote, United States",
    "Remote US",
    "US-Based",
    "US based",
)

DEFAULT_INTERN_REGEX = r"\b(internship|intern|co[-\s]?op)\b"
DEFAULT_GRAD_REGEX = (
    r"\b(ncg|new\s+college\s+grad|new\s+grad|university\s+graduate|gred|early\s+career|"
    r"rotational\s+graduate)\b"
)

# If the location/title names another country, we reject unless US also appears there
# (comma + US state counts as US). Stops multinational boilerplate in the body from
# letting "Munich, Germany" or "Tokyo, Japan" listings through.
NON_US_COUNTRY_RE = re.compile(
    r"\b("
    r"Afghanistan|Albania|Algeria|Argentina|Armenia|Australia|Austria|Azerbaijan|"
    r"Bahrain|Bangladesh|Belarus|Belgium|Brazil|Brunei|Bulgaria|Cambodia|Canada|Chile|"
    r"China|Colombia|Costa Rica|Croatia|Cyprus|Czech Republic|Czechia|Denmark|Ecuador|"
    r"Egypt|Estonia|Ethiopia|Finland|France|Germany|Ghana|Greece|Hungary|Iceland|India|"
    r"Indonesia|Iran|Iraq|Ireland|Israel|Italy|Japan|Jordan|Kazakhstan|Kenya|Kuwait|"
    r"Latvia|Lebanon|Lithuania|Luxembourg|Malaysia|Malta|Morocco|Myanmar|Nepal|"
    r"Netherlands|New Zealand|Nigeria|Norway|Oman|Pakistan|Panama|Peru|Philippines|Poland|"
    r"Portugal|Qatar|Romania|Russia|Saudi Arabia|Serbia|Singapore|Slovakia|Slovenia|"
    r"South Africa|South Korea|Spain|Sri Lanka|Sweden|Switzerland|Taiwan|Thailand|"
    r"Turkey|Ukraine|United Arab Emirates|UAE|Vietnam|Viet Nam|"
    r"(?<!New\s)Mexico(?!\s+Rico)|"
    r"United Kingdom|UK|England|Scotland|Wales|Northern Ireland|"
    r"Hong Kong|Macau"
    r")\b",
    re.IGNORECASE,
)

# Omit IN: job boards often use "Bengaluru, IN" for India, not Indiana.
_US_STATE_CODES = (
    "AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IA|KS|KY|LA|ME|MD|MA|MI|MN|MS|MO|MT|"
    "NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|DC"
)
# Typical US job board format: "City, ST" or "…; City, ST"
US_STATE_COMMA_RE = re.compile(rf",\s*({_US_STATE_CODES})\b", re.IGNORECASE)

# Canadian province/territory on job boards: "Waterloo, ON"
CA_PROVINCE_COMMA_RE = re.compile(
    r",\s*(ON|BC|AB|SK|MB|QC|NS|NB|NL|PE|YT|NT|NU)\b",
    re.IGNORECASE,
)

CANADIAN_METRO_RE = re.compile(
    r"\b("
    r"Toronto|Montreal|Montréal|Vancouver|Calgary|Ottawa|Edmonton|Winnipeg|Mississauga|"
    r"Halifax|Hamilton|Quebec City|Quebec|Gatineau|Saskatoon|Regina|"
    r"Kitchener|Waterloo|Barrie|Oshawa|Windsor"
    r")\b",
    re.IGNORECASE,
)


def _comma_state_signals_us(text: str) -> bool:
    for m in US_STATE_COMMA_RE.finditer(text):
        code = m.group(1).upper()
        if code == "CA":
            ctx = text[max(0, m.start() - 120) : m.start()]
            if CANADIAN_METRO_RE.search(ctx):
                continue
        return True
    return False


def _compile_regex(pattern: str | None, fallback: str) -> re.Pattern[str]:
    src = (pattern or "").strip() or fallback
    try:
        return re.compile(src, re.IGNORECASE)
    except re.error:
        return re.compile(fallback, re.IGNORECASE)


def build_scan_preferences(raw: dict[str, Any]) -> tuple[bool, tuple[str, ...], re.Pattern[str], re.Pattern[str]]:
    """
    Returns (united_states_only, us_indicators, intern_re, grad_re).
    Role priority: intern (0) < grad/ncg (1) < other (2) for sorting.
    """
    us_only = bool(raw.get("united_states_only", True))
    inds = raw.get("us_location_indicators")
    if isinstance(inds, list) and inds:
        indicators = tuple(str(x).strip() for x in inds if isinstance(x, str) and str(x).strip())
    else:
        indicators = DEFAULT_US_LOCATION_INDICATORS
    rs = raw.get("role_sort")
    intern_pat: str | None = None
    grad_pat: str | None = None
    if isinstance(rs, dict):
        intern_pat = rs.get("intern_regex") if isinstance(rs.get("intern_regex"), str) else None
        grad_pat = rs.get("grad_regex") if isinstance(rs.get("grad_regex"), str) else None
    intern_re = _compile_regex(intern_pat, DEFAULT_INTERN_REGEX)
    grad_re = _compile_regex(grad_pat, DEFAULT_GRAD_REGEX)
    return us_only, indicators, intern_re, grad_re


def _job_geo_text(job: NormalizedJob) -> str:
    return f"{job.location or ''}\n{job.title}\n{job.body}"


def _haystack_has_us_indicator(hay: str, indicators: tuple[str, ...]) -> bool:
    h = hay.lower()
    for ind in indicators:
        s = ind.lower().strip()
        if not s:
            continue
        if len(s) <= 4:
            if re.search(rf"(?<![a-z0-9]){re.escape(s)}(?![a-z0-9])", h, re.IGNORECASE):
                return True
        else:
            if s in h:
                return True
    return False


def _field_signals_us(text: str, indicators: tuple[str, ...]) -> bool:
    """US in this field: explicit phrases or common `City, ST` pattern."""
    if not text.strip():
        return False
    if _haystack_has_us_indicator(text, indicators):
        return True
    return _comma_state_signals_us(text)


def _field_contradicts_us_only(text: str, indicators: tuple[str, ...]) -> bool:
    """Location/title names another country (or Canada) without a US signal in the same text."""
    if not text.strip():
        return False
    if CANADIAN_METRO_RE.search(text) or CA_PROVINCE_COMMA_RE.search(text):
        if not _field_signals_us(text, indicators):
            return True
    if NON_US_COUNTRY_RE.search(text):
        return not _field_signals_us(text, indicators)
    return False


_VAGUE_MULTI_LOCATION_RE = re.compile(
    r"^(?P<n>\d+)\s+locations?$|^(multiple|various|several)\s+locations?$",
    re.IGNORECASE,
)


def _workday_location_is_vague_us_unknown(loc: str) -> bool:
    """Workday list API often returns '3 Locations' with no country text — not a US contradiction."""
    s = loc.strip()
    if not s:
        return False
    if _VAGUE_MULTI_LOCATION_RE.match(s):
        return True
    sl = s.casefold()
    return sl in ("global", "worldwide", "multiple sites", "various sites")


def job_is_united_states(job: NormalizedJob, indicators: tuple[str, ...]) -> bool:
    loc = (job.location or "").strip()
    title = (job.title or "").strip()
    if _field_contradicts_us_only(loc, indicators) or _field_contradicts_us_only(title, indicators):
        return False
    if _field_signals_us(loc, indicators) or _field_signals_us(title, indicators):
        return True
    if _haystack_has_us_indicator(_job_geo_text(job), indicators):
        return True
    # Multi-site placeholder only: allow if title does not name a non-US region (avoids dropping
    # US-heavy reqs where Workday omits "US, CA, …" on the list card).
    if _workday_location_is_vague_us_unknown(loc) and not _field_contradicts_us_only(title, indicators):
        return True
    return False


def job_role_priority(job: NormalizedJob, intern_re: re.Pattern[str], grad_re: re.Pattern[str]) -> int:
    """0 = intern/co-op, 1 = new grad / NCG / early career, 2 = other."""
    text = _job_geo_text(job)
    if intern_re.search(text):
        return 0
    if grad_re.search(text):
        return 1
    return 2


_WORKDAY_POSTED_DAYS_AGO_RE = re.compile(r"posted\s+(\d+)\s+days?\s+ago", re.IGNORECASE)
_WORKDAY_POSTED_ON_MDY_RE = re.compile(
    r"posted\s+on\s+(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})",
    re.IGNORECASE,
)


def _posted_ts_for_sort(posted_at: str | None) -> float:
    """
    Monotonic "freshness" timestamp for sorting: larger = more recently posted.
    Handles ISO dates, YYYY-MM-DD, and Workday UI strings (Posted Today, Posted N Days Ago, etc.).
    """
    if not posted_at or not str(posted_at).strip():
        return 0.0
    s = " ".join(str(posted_at).split()).strip()
    try:
        iso_s = s
        if iso_s.endswith("Z"):
            iso_s = iso_s[:-1] + "+00:00"
        return datetime.fromisoformat(iso_s.replace("Z", "+00:00")).timestamp()
    except ValueError:
        pass
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            d = datetime.strptime(s[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
            return d.timestamp()
        except ValueError:
            pass
    sl = s.lower()
    now = datetime.now(timezone.utc)
    # Workday list API: "Posted Today", or UI blocks like "Posted on" + "Posted Today" (whitespace-normalized).
    if re.search(r"\bposted\s+today\b", sl):
        return now.timestamp()
    if re.search(r"\bposted\s+just\s+now\b", sl) or (
        "just now" in sl and "posted" in sl
    ):
        return now.timestamp()
    if "yesterday" in sl and "posted" in sl:
        return (now - timedelta(days=1)).timestamp()
    mh = re.search(r"posted\s+(\d+)\s+hours?\s+ago", s, re.IGNORECASE)
    if mh:
        h = int(mh.group(1))
        return (now - timedelta(seconds=h * 3600)).timestamp()
    m = _WORKDAY_POSTED_DAYS_AGO_RE.search(s)
    if m:
        return (now - timedelta(days=int(m.group(1)))).timestamp()
    mw = re.search(r"posted\s+(\d+)\s+weeks?\s+ago", s, re.IGNORECASE)
    if mw:
        return (now - timedelta(days=7 * int(mw.group(1)))).timestamp()
    mm = re.search(r"posted\s+(\d+)\s+months?\s+ago", s, re.IGNORECASE)
    if mm:
        return (now - timedelta(days=30 * int(mm.group(1)))).timestamp()
    mon = re.search(
        r"posted\s+on\s+([a-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})",
        sl,
        re.IGNORECASE,
    )
    if mon:
        mword = mon.group(1).title()
        dnum, ynum = int(mon.group(2)), mon.group(3)
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                d = datetime.strptime(f"{mword} {dnum} {ynum}", fmt).replace(
                    tzinfo=timezone.utc
                )
                return d.timestamp()
            except ValueError:
                continue
    mdy = _WORKDAY_POSTED_ON_MDY_RE.search(s)
    if mdy:
        mo, da, yr = int(mdy.group(1)), int(mdy.group(2)), int(mdy.group(3))
        if yr < 100:
            yr += 2000
        try:
            d = datetime(yr, mo, da, tzinfo=timezone.utc)
            return d.timestamp()
        except ValueError:
            pass
    return 0.0


def _posted_recency_sort_key(posted_at: str | None) -> float:
    """Ascending sort: lower = fresher board post; unknown dates last."""
    ts = _posted_ts_for_sort(posted_at)
    if ts <= 0:
        return 9.0e15
    return -float(ts)


def fetch_all(companies: list[dict[str, Any]]) -> list[NormalizedJob]:
    all_jobs: list[NormalizedJob] = []
    for row in companies:
        if row.get("enabled") is False:
            continue
        name = row.get("name") or "Unknown"
        ats = (row.get("ats") or "").lower().strip()
        try:
            if ats == "greenhouse":
                board = row.get("board")
                if not board:
                    continue
                all_jobs.extend(fetch_greenhouse(str(board), name))
            elif ats == "lever":
                site = row.get("site")
                if not site:
                    continue
                all_jobs.extend(fetch_lever(str(site), name))
            elif ats == "workday":
                tenant_url = row.get("tenant_url")
                portal_base = row.get("portal_base")
                if not tenant_url or not portal_base:
                    print(f"[skip] {name}: workday needs tenant_url and portal_base", file=sys.stderr)
                    continue
                er = row.get("workday_enrich_job_details")
                enrich = er is True or (isinstance(er, str) and er.strip().lower() in ("true", "1", "yes"))
                try:
                    emax = int(row.get("workday_enrich_max") or 0)
                except (TypeError, ValueError):
                    emax = 0
                all_jobs.extend(
                    fetch_workday(
                        str(tenant_url),
                        str(portal_base),
                        name,
                        row.get("workday_alternate"),
                        enrich_job_details=enrich,
                        enrich_max=emax if enrich else 0,
                    )
                )
            elif ats == "eightfold":
                ef_host = row.get("eightfold_host")
                ef_domain = row.get("eightfold_domain")
                if not ef_host or not ef_domain:
                    print(f"[skip] {name}: eightfold needs eightfold_host and eightfold_domain", file=sys.stderr)
                    continue
                all_jobs.extend(fetch_eightfold(str(ef_host), str(ef_domain), name))
            elif ats == "ashby":
                org = row.get("org")
                if not org:
                    print(f"[skip] {name}: ashby needs org", file=sys.stderr)
                    continue
                all_jobs.extend(fetch_ashby(str(org), name))
            elif ats == "phenom":
                listing = row.get("listing_url")
                if not listing:
                    print(f"[skip] {name}: phenom needs listing_url", file=sys.stderr)
                    continue
                max_pg = int(row.get("phenom_max_pages") or 15)
                pq = str(row.get("phenom_page_query") or "pg").strip() or "pg"
                all_jobs.extend(fetch_phenom(str(listing), name, max_pages=max_pg, page_query=pq))
            elif ats == "jobvite":
                slug = row.get("jobvite_slug")
                if not slug:
                    print(f"[skip] {name}: jobvite needs jobvite_slug", file=sys.stderr)
                    continue
                all_jobs.extend(fetch_jobvite(str(slug), name))
            elif ats == "mirafra":
                murl = row.get("mirafra_careers_url") or "https://mirafra.com/career/"
                all_jobs.extend(fetch_mirafra(str(murl), name))
            elif ats == "smartrecruiters":
                sr_id = row.get("smartrecruiters_company")
                if not sr_id:
                    print(f"[skip] {name}: smartrecruiters needs smartrecruiters_company", file=sys.stderr)
                    continue
                all_jobs.extend(fetch_smartrecruiters(str(sr_id), name))
            elif ats == "jobspy":
                queries = row.get("jobspy_queries")
                if not isinstance(queries, list) or not queries:
                    print(f"[skip] {name}: jobspy needs jobspy_queries list", file=sys.stderr)
                    continue
                qlist = [str(x).strip() for x in queries if isinstance(x, str) and str(x).strip()]
                if not qlist:
                    continue
                all_jobs.extend(fetch_jobspy(qlist, name))
            else:
                print(f"[skip] {name}: unknown ats '{ats}'", file=sys.stderr)
        except requests.HTTPError as e:
            print(f"[error] {name} ({ats}): HTTP {e.response.status_code}", file=sys.stderr)
        except requests.RequestException as e:
            print(f"[error] {name} ({ats}): {e}", file=sys.stderr)
        time.sleep(0.15)
    return all_jobs


def upsert_and_classify(
    conn: sqlite3.Connection,
    job: NormalizedJob,
    score: float,
    tracks: list[str],
    now_iso: str,
) -> tuple[bool, bool]:
    """
    Returns (is_new_row, needs_your_attention).
    needs_your_attention = new this run OR (exists and not applied).
    """
    text_blob = f"{job.title}\n{job.body}"
    cur = conn.cursor()
    cur.execute(
        "SELECT applied, first_seen_at FROM jobs WHERE source_key = ? AND external_id = ?",
        (job.source_key, job.external_id),
    )
    row = cur.fetchone()
    if row is None:
        try:
            conn.execute(
                """
                INSERT INTO jobs (
                    source_key, external_id, title, company_name, url, location,
                    posted_at, first_seen_at, last_seen_at, applied, relevance_score, tracks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                """,
                (
                    job.source_key,
                    job.external_id,
                    job.title,
                    job.company_name,
                    job.url,
                    job.location,
                    job.posted_at,
                    now_iso,
                    now_iso,
                    score,
                    json.dumps(tracks),
                ),
            )
        except sqlite3.IntegrityError as e:
            # Same job URL can appear twice (e.g. two JobSpy queries or ATS + aggregator) with
            # different (source_key, external_id); PRIMARY KEY differs but url UNIQUE forbids a second row.
            if "jobs.url" not in str(e):
                raise
            cur.execute("SELECT applied FROM jobs WHERE url = ?", (job.url,))
            dup = cur.fetchone()
            if dup is None:
                raise
            applied_dup = int(dup[0])
            conn.execute(
                """
                UPDATE jobs SET last_seen_at = ?, relevance_score = ?, tracks = ?
                WHERE url = ?
                """,
                (now_iso, score, json.dumps(tracks), job.url),
            )
            return False, applied_dup == 0
        return True, True
    applied = int(row[0])
    conn.execute(
        """
        UPDATE jobs SET
            title = ?, company_name = ?, url = ?, location = ?, posted_at = ?,
            last_seen_at = ?, relevance_score = ?, tracks = ?
        WHERE source_key = ? AND external_id = ?
        """,
        (
            job.title,
            job.company_name,
            job.url,
            job.location,
            job.posted_at,
            now_iso,
            score,
            json.dumps(tracks),
            job.source_key,
            job.external_id,
        ),
    )
    return False, applied == 0


def mac_notify(title: str, subtitle: str, message: str) -> None:
    if sys.platform != "darwin":
        return

    def esc(s: str) -> str:
        return s.replace("\\", "\\\\").replace('"', '\\"')

    script = (
        f'display notification "{esc(message)}" '
        f'with title "{esc(title)}" '
        f'subtitle "{esc(subtitle)}" '
        f'sound name "Glass"'
    )
    subprocess.run(["osascript", "-e", script], check=False, capture_output=True)


def _markdown_link_label_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("[", "\\[").replace("]", "\\]")


# --- daily_logs markdown (dedupe across all *.md, prepend new jobs to today's file) ---

_BOT_JOB_TOKEN_RE = re.compile(r"<!--\s*bot-job:([A-Za-z0-9_-]+)\s*-->")
_LOG_JOB_URL_RE = re.compile(r"^##\s+\[[^\]]*\]\(([^)]+)\)\s*$", re.MULTILINE)


def _job_log_token(job: NormalizedJob) -> str:
    raw = f"{job.source_key}\0{job.external_id}".encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def collect_logged_job_keys(logs_dir: Path) -> set[str]:
    """
    Keys already written to any daily log: Greenhouse-style tokens plus url:<job_url>
    for legacy entries without <!-- bot-job:... --> markers.
    """
    keys: set[str] = set()
    if not logs_dir.is_dir():
        return keys
    for path in sorted(logs_dir.glob("*.md")):
        try:
            text = path.read_text(encoding="utf-8")
        except OSError:
            continue
        for m in _BOT_JOB_TOKEN_RE.finditer(text):
            keys.add(m.group(1))
        for m in _LOG_JOB_URL_RE.finditer(text):
            keys.add(f"url:{m.group(1).strip()}")
    return keys


def _strip_daily_log_header_and_stamps(text: str) -> str:
    """Remove title line and *...* status lines; return job markdown body."""
    if not text.strip():
        return ""
    lines = text.lstrip("\ufeff").splitlines()
    i = 0
    if i < len(lines) and lines[i].startswith("# New Semiconductor Roles"):
        i += 1
    while i < len(lines) and not lines[i].strip():
        i += 1
    while i < len(lines) and lines[i].lstrip().startswith("*"):
        i += 1
        while i < len(lines) and not lines[i].strip():
            i += 1
    return "\n".join(lines[i:]).lstrip("\n")


def _format_job_log_block(job: NormalizedJob, score: float, matched: list[str]) -> str:
    tok = _job_log_token(job)
    loc = str(job.location or "").strip() or "—"
    track_str = ", ".join(matched) if matched else "—"
    label = _markdown_link_label_escape(f"{job.company_name} — {job.title}")
    posted_line = ""
    pa = str(job.posted_at or "").strip()
    if pa:
        posted_line = f"\n- **Posted:** {pa}"
    extra_urls = [
        ln.strip()
        for ln in (job.body or "").splitlines()
        if ln.strip().startswith(("http://", "https://"))
    ]
    extra_block = ""
    if extra_urls:
        lines = [f"- **Additional Indeed URLs ({len(extra_urls)}):**"]
        for u in extra_urls:
            lines.append(f"  - [{_markdown_link_label_escape(u)}]({u})")
        extra_block = "\n" + "\n".join(lines)
    return (
        f"<!-- bot-job:{tok} -->\n"
        f"## [{label}]({job.url})\n\n"
        f"- **Company:** {job.company_name}\n"
        f"- **Title:** {job.title}\n"
        f"- **Location:** {loc}"
        f"{posted_line}\n"
        f"- **Score:** {score:.2f}\n"
        f"- **Tracks:** {track_str}"
        f"{extra_block}\n"
    )


def write_daily_roles_log(
    logs_dir: Path,
    date_str: str,
    new_job_rows: list[tuple[NormalizedJob, float, list[str]]],
) -> None:
    """
    Update daily_logs/{date}.md: recreate if missing; prepend jobs not already
    mentioned in any *.md in logs_dir (by bot-job token or job URL).
    """
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / f"{date_str}.md"

    logged = collect_logged_job_keys(logs_dir)
    to_prepend: list[tuple[NormalizedJob, float, list[str]]] = []
    for job, score, matched in new_job_rows:
        tok = _job_log_token(job)
        url_key = f"url:{job.url}"
        if tok in logged or url_key in logged:
            continue
        to_prepend.append((job, score, matched))
        logged.add(tok)
        logged.add(url_key)

    existing_raw = ""
    if log_path.exists():
        try:
            existing_raw = log_path.read_text(encoding="utf-8")
        except OSError:
            existing_raw = ""

    remainder = _strip_daily_log_header_and_stamps(existing_raw)
    header = f"# New Semiconductor Roles - {date_str}\n\n"
    now_s = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    if to_prepend:
        stamp = f"*Updated: {now_s}* — {len(to_prepend)} new listing(s) prepended (not present in other log files).\n\n"
    else:
        stamp = (
            f"*Updated: {now_s}* — no new listings this run, or every new job was already "
            f"recorded in a log under {logs_dir.name}/.\n\n"
        )

    new_blocks = [_format_job_log_block(j, s, m) for j, s, m in to_prepend]
    parts: list[str] = [header, stamp]
    if new_blocks:
        parts.append("\n\n".join(new_blocks))
        if remainder.strip():
            parts.append("\n\n")
    if remainder.strip():
        parts.append(remainder.strip())
        parts.append("\n")

    try:
        log_path.write_text("".join(parts), encoding="utf-8")
    except OSError as e:
        print(f"[warn] could not write {log_path}: {e}", file=sys.stderr)


def _excel_job_id_display(external_id: str, url: str) -> str:
    m = _EXCEL_JOB_ID_JR_RE.search(url or "")
    if m:
        return m.group(0).upper().replace("_", "-")
    eid = (external_id or "").strip()
    if not eid:
        u = url or ""
        return u[-120:] if len(u) > 120 else u
    if "/" in eid:
        return eid.split("/")[-1][:120]
    return eid[:120]


def _excel_parse_first_seen(iso_s: str | None) -> datetime | None:
    if not iso_s or not str(iso_s).strip():
        return None
    s = str(iso_s).strip()
    try:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        # openpyxl rejects tz-aware datetimes; store as naive UTC wall time.
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except ValueError:
        return None


def export_jobs_excel(
    conn: sqlite3.Connection,
    intern_re: re.Pattern[str],
    out_path: Path | None = None,
) -> Path:
    """
    Write jobs_tracker.xlsx with two sheets (co-op/intern vs full time).
    Sort: not applied first, then applied; within each group, most recent job-board posted date
    first (Workday-style "Posted Today" / "Posted N Days Ago", etc.), then first_seen_at as tiebreaker.
    Columns end with Applied (Excel boolean FALSE/TRUE) and URL (clickable hyperlink).
    Editing the file does not update the DB.
    """
    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.worksheet import Worksheet

    out = Path(out_path or EXCEL_TRACKER_PATH)
    headers = [
        "Job ID",
        "Company name",
        "Job name",
        "Locations",
        "Posted (job board)",
        "Date collected",
        "Applied",
        "URL",
    ]
    link_font = Font(color="0563C1", underline="single")

    cur = conn.cursor()
    cur.execute(
        """
        SELECT external_id, company_name, title, location, posted_at, first_seen_at, applied, url
        FROM jobs
        """
    )
    raw = cur.fetchall()

    rows: list[tuple[str, str, str, str | None, str, str, int, str]] = []
    for external_id, company_name, title, location, posted_at, first_seen_at, applied, url in raw:
        rows.append(
            (
                str(external_id),
                str(company_name),
                str(title),
                location,
                str(posted_at).strip() if posted_at else "",
                str(first_seen_at) if first_seen_at else "",
                int(applied),
                str(url),
            )
        )

    def is_intern_coop(r: tuple[str, str, str, str | None, str, str, int, str]) -> bool:
        return bool(intern_re.search(r[2] or ""))

    intern_rows = [r for r in rows if is_intern_coop(r)]
    full_rows = [r for r in rows if not is_intern_coop(r)]

    def excel_sort_key(r: tuple[str, str, str, str | None, str, str, int, str]) -> tuple[int, float, float]:
        _eid, _company, _title, _loc, po, fs, ap, _url = r
        fs_ts = _posted_ts_for_sort(fs)
        fs_key = -fs_ts if fs_ts > 0 else 0.0
        return (ap, _posted_recency_sort_key(po if po else None), fs_key)

    intern_rows.sort(key=excel_sort_key)
    full_rows.sort(key=excel_sort_key)

    wb = Workbook()
    ws_intern = wb.active
    if ws_intern is None:
        raise RuntimeError("openpyxl workbook has no active worksheet")
    ws_intern.title = "Co-op & Intern"
    ws_full = wb.create_sheet("Full time")

    def write_sheet(ws: Worksheet, data: list[tuple[str, str, str, str | None, str, str, int, str]]) -> None:
        ws.append(list(headers))
        for external_id, company_nm, title, location, posted_board, first_seen_at, applied, url in data:
            jid = _excel_job_id_display(external_id, url)
            dt = _excel_parse_first_seen(first_seen_at)
            loc_s = str(location or "").strip()
            posted_s = str(posted_board or "").strip()
            applied_bool = bool(applied)
            ws.append(
                [
                    jid,
                    company_nm,
                    title,
                    loc_s,
                    posted_s,
                    dt if dt is not None else (first_seen_at or ""),
                    applied_bool,
                    url,
                ]
            )
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        for col_letter, width in (
            ("A", 16),
            ("B", 26),
            ("C", 48),
            ("D", 28),
            ("E", 24),
            ("F", 20),
            ("G", 10),
            ("H", 54),
        ):
            ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
        top_align = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
            for cell in row:
                cell.alignment = top_align
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            for cell in row:
                if not isinstance(cell, Cell):
                    continue
                u = cell.value
                if isinstance(u, str) and u.startswith(("http://", "https://")):
                    cell.hyperlink = u
                    cell.font = link_font
                    cell.alignment = top_align

    write_sheet(ws_intern, intern_rows)
    write_sheet(ws_full, full_rows)

    wb.save(out)
    return out


def cmd_scan(args: argparse.Namespace) -> int:
    companies_cfg = load_yaml(COMPANIES_PATH)
    kw_cfg = load_yaml(KEYWORDS_PATH)
    tracks, token_groups, min_score = compile_keyword_config(kw_cfg)
    skip_perf_companies = skip_performance_track_company_set(kw_cfg)
    us_only, us_inds, intern_re, grad_re = build_scan_preferences(kw_cfg)
    exclude_title = compile_exclude_title_patterns(kw_cfg)

    companies = companies_cfg.get("companies") or []
    if not companies:
        print("No companies in companies.yaml", file=sys.stderr)
        return 1

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    now = datetime.now(timezone.utc).isoformat()

    raw_jobs = fetch_all(companies)
    new_jobs: list[tuple[NormalizedJob, float, list[str]]] = []
    unapplied_jobs: list[tuple[NormalizedJob, float, list[str], bool]] = []

    for job in raw_jobs:
        blob = f"{job.title}\n{job.body}"
        score, matched = score_job(
            blob,
            tracks,
            token_groups=token_groups,
            company_name=job.company_name,
            skip_performance_companies=skip_perf_companies,
        )
        if score < min_score:
            continue
        if job_title_excluded(job.title, exclude_title):
            continue
        if us_only and not job_is_united_states(job, us_inds):
            continue
        is_new, attention = upsert_and_classify(conn, job, score, matched, now)
        if is_new:
            new_jobs.append((job, score, matched))
        if attention:
            unapplied_jobs.append((job, score, matched, is_new))

    def _new_jobs_recency_key(row: tuple[NormalizedJob, float, list[str]]) -> tuple[float, str, str]:
        j, _, __ = row
        return (
            _posted_recency_sort_key(j.posted_at),
            j.company_name.lower(),
            j.title.lower(),
        )

    new_jobs.sort(key=_new_jobs_recency_key)

    today_str = datetime.now().strftime("%Y-%m-%d")
    write_daily_roles_log(LOGS_DIR, today_str, new_jobs)

    conn.commit()

    try:
        export_jobs_excel(conn, intern_re, EXCEL_TRACKER_PATH)
    except Exception as e:
        print(f"[warn] could not write {EXCEL_TRACKER_PATH.name}: {e}", file=sys.stderr)

    def _scan_sort_key(
        item: tuple[NormalizedJob, float, list[str], bool],
    ) -> tuple[float, int, int, str, str]:
        job, _score, _matched, is_new = item
        pri = job_role_priority(job, intern_re, grad_re)
        new_rank = 0 if is_new else 1
        return (
            _posted_recency_sort_key(job.posted_at),
            pri,
            new_rank,
            job.company_name.lower(),
            job.title.lower(),
        )

    lines: list[str] = []
    for job, score, matched, is_new in sorted(unapplied_jobs, key=_scan_sort_key):
        flag = "NEW  " if is_new else "OPEN "
        track_str = ", ".join(matched)
        loc = f" | {job.location}" if job.location else ""
        posted_hint = ""
        pa = str(job.posted_at or "").strip()
        if pa:
            posted_hint = f"     posted: {pa}\n"
        chunk = (
            f"{flag} [{job.company_name}]{loc}\n"
            f"     {job.title}\n"
            f"{posted_hint}"
            f"     score={score:.2f}  tracks: {track_str}\n"
            f"     {job.url}\n"
        )
        for ln in (job.body or "").splitlines():
            u = ln.strip()
            if u.startswith(("http://", "https://")):
                chunk += f"     also: {u}\n"
        lines.append(chunk)

    if not lines:
        print("No matching unapplied roles right now (or all sources failed).")
        conn.close()
        return 0

    report = "\n".join(lines)
    print(report)

    if getattr(args, "notify", False) and new_jobs:
        titles_preview = " · ".join(
            f"{j.company_name} — {j.title}" for j, _, _ in new_jobs[:2]
        )[:100]
        mac_notify(
            "🔔 New Semiconductor Roles",
            f"{len(new_jobs)} new role{'s' if len(new_jobs) > 1 else ''} found",
            titles_preview,
        )

    conn.close()
    return 0


def cmd_daemon(args: argparse.Namespace) -> int:
    interval_mins = args.interval
    print(f"[daemon] Started — scanning every {interval_mins} minutes.")
    print(f"[daemon] Press Ctrl+C to stop.\n")

    scan_count = 0
    while True:
        scan_count += 1
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[daemon] Scan #{scan_count} at {now_str}")
        try:
            cmd_scan(args)
        except Exception as e:
            print(f"[daemon] Scan error (will retry next interval): {e}", file=sys.stderr)
        next_run = datetime.now() + timedelta(minutes=interval_mins)
        print(
            f"[daemon] Next scan at {next_run.strftime('%H:%M:%S')} "
            f"(in {interval_mins} minutes)\n"
        )
        time.sleep(interval_mins * 60)


def cmd_mark_applied(args: argparse.Namespace) -> int:
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    cur = conn.cursor()
    n = 0
    if args.url:
        cur.execute("UPDATE jobs SET applied = 1 WHERE url = ?", (args.url.strip(),))
        n = cur.rowcount
    elif args.source and args.ext_id:
        cur.execute(
            "UPDATE jobs SET applied = 1 WHERE source_key = ? AND external_id = ?",
            (args.source.strip(), args.ext_id.strip()),
        )
        n = cur.rowcount
    else:
        print("Provide --url or both --source-key and --external-id", file=sys.stderr)
        conn.close()
        return 1
    conn.commit()
    if n == 0:
        conn.close()
        print("No row updated; check URL or ids.", file=sys.stderr)
        return 2
    try:
        kw_cfg = load_yaml(KEYWORDS_PATH)
        _, _, intern_re, _ = build_scan_preferences(kw_cfg)
        export_jobs_excel(conn, intern_re, EXCEL_TRACKER_PATH)
    except Exception as e:
        print(f"[warn] could not refresh {EXCEL_TRACKER_PATH.name}: {e}", file=sys.stderr)
    conn.close()
    print(f"Marked applied ({n} row).")
    return 0


def cmd_export_excel(_args: argparse.Namespace) -> int:
    kw_cfg = load_yaml(KEYWORDS_PATH)
    _, _, intern_re, _ = build_scan_preferences(kw_cfg)
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    try:
        path = export_jobs_excel(conn, intern_re, EXCEL_TRACKER_PATH)
    except Exception as e:
        print(f"Excel export failed: {e}", file=sys.stderr)
        conn.close()
        return 1
    conn.close()
    print(f"Wrote {path}")
    return 0


def cmd_list_unapplied(_args: argparse.Namespace) -> int:
    kw_cfg = load_yaml(KEYWORDS_PATH)
    us_only, us_inds, intern_re, grad_re = build_scan_preferences(kw_cfg)
    exclude_title = compile_exclude_title_patterns(kw_cfg)

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT source_key, external_id, company_name, title, url, location, posted_at,
               first_seen_at, tracks, relevance_score
        FROM jobs WHERE applied = 0
        """
    )
    rows = cur.fetchall()
    conn.close()
    if not rows:
        print("No unapplied jobs in database.")
        return 0

    decorated: list[tuple[NormalizedJob, str, str, float]] = []
    for sk, eid, company, title, url, loc, posted, first_seen, tracks_json, score in rows:
        job = NormalizedJob(
            source_key=sk,
            external_id=eid,
            title=title,
            company_name=company,
            url=url,
            location=loc,
            posted_at=posted,
            body="",
        )
        if us_only and not job_is_united_states(job, us_inds):
            continue
        if job_title_excluded(job.title, exclude_title):
            continue
        decorated.append((job, first_seen, tracks_json, float(score)))

    def _list_sort_key(
        t: tuple[NormalizedJob, str, str, float],
    ) -> tuple[float, int, float, str, str]:
        job, first_seen, _tracks_json, _score = t
        pri = job_role_priority(job, intern_re, grad_re)
        fs_ts = _posted_ts_for_sort(first_seen)
        fs_key = -fs_ts if fs_ts > 0 else 0.0
        return (
            _posted_recency_sort_key(job.posted_at),
            pri,
            fs_key,
            job.company_name.lower(),
            job.title.lower(),
        )

    decorated.sort(key=_list_sort_key)

    for job, first_seen, tracks_json, score in decorated:
        loc = f" | {job.location}" if job.location else ""
        board_posted = ""
        pa = str(job.posted_at or "").strip()
        if pa:
            board_posted = f"  board posted: {pa}\n"
        print(
            f"- [{job.company_name}]{loc} {job.title}\n"
            f"{board_posted}"
            f"  {job.url}\n"
            f"  since {first_seen}  score={score}  {tracks_json}\n"
        )
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Semiconductor design/DV/perf modeling job scanner.")
    sub = p.add_subparsers(dest="command")

    ps = sub.add_parser("scan", help="Fetch boards, match keywords, show new / unapplied.")
    ps.add_argument("--notify", action="store_true", help="macOS notification when new jobs appear.")
    ps.set_defaults(func=cmd_scan)

    pd = sub.add_parser("daemon", help="Run scan on repeat. Ctrl+C to stop.")
    pd.add_argument(
        "--interval",
        type=int,
        default=360,
        help="Minutes between scans (default: 360 = 6 hours).",
    )
    pd.add_argument(
        "--notify",
        action="store_true",
        help="macOS desktop popup when new jobs appear.",
    )
    pd.set_defaults(func=cmd_daemon)

    pm = sub.add_parser("mark-applied", help="Mark a job as applied (stops highlighting).")
    pm.add_argument("--url", help="Exact job URL as printed by scan.")
    pm.add_argument("--source-key", dest="source", help="e.g. greenhouse:nvidia")
    pm.add_argument("--external-id", dest="ext_id", help="External job id from ATS.")
    pm.set_defaults(func=cmd_mark_applied)

    pl = sub.add_parser("list-unapplied", help="List all unapplied jobs in DB.")
    pl.set_defaults(func=cmd_list_unapplied)

    pe = sub.add_parser(
        "export-excel",
        help=f"Write {EXCEL_TRACKER_PATH.name} (co-op/intern vs full time sheets, sorted like scan).",
    )
    pe.set_defaults(func=cmd_export_excel)

    return p


def main() -> None:
    parser = build_parser()
    argv = sys.argv[1:]
    if not argv:
        argv = ["scan"]
    elif argv[0].startswith("-"):
        argv = ["scan"] + argv
    args = parser.parse_args(argv)
    if not getattr(args, "func", None):
        parser.print_help()
        raise SystemExit(2)
    raise SystemExit(args.func(args))


if __name__ == "__main__":
    main()
